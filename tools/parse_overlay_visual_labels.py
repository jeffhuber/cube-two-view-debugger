#!/usr/bin/env python3
"""Parse the May 18 overlay-tool human-feedback xlsx into a structured
per-(set, side, slot) label artifact at
`tests/fixtures/overlay_visual_labels.json`.

Input xlsx schema (single sheet, two-row header):
    row 1: section labels — "A quads", "A rectified", "B quads", "B rectified"
    row 2: per-column headers:
        Image Set | red outline | green outline | blue outline |
        U face rectified (Slot = U) | R face rectified (Slot = R) |
        F face rectified (Slot = F) | purple outline | orange outline |
        yellow outline | D face rectified (Slot = D) |
        L face rectified (Slot = L) | B face rectified (Slot = B)

    rows 3+: one row per image set:
        17 | good | good | bad | src = U; good | src = R; good |
             src = B; bad | bad | bad | bad | src = D; bad |
             src = B; bad | src = B; bad

Output JSON schema:
    {
      "source": "/private/tmp/hybrid_overlays/overlay tool visual feedback  (1) copy.xlsx",
      "captured_at": "2026-05-18",  # date of the overlay generation
      "pipeline_version": "pre_PR_160",  # what the overlays were generated from
      "color_to_slot": {
        "A": {"red": "U", "green": "R", "blue": "F"},
        "B": {"purple": "D", "orange": "L", "yellow": "B"}
      },
      "pairs": [
        {
          "set_id": "17",
          "quads": {
            "A": {"U": "good", "R": "good", "F": "bad"},
            "B": {"D": "bad", "L": "bad", "B": "bad"}
          },
          "rectified": {
            "A": {
              "U": {"src_face": "U", "verdict": "good"},
              "R": {"src_face": "R", "verdict": "good"},
              "F": {"src_face": "B", "verdict": "bad"}
            },
            "B": {
              "D": {"src_face": "D", "verdict": "bad"},
              "L": {"src_face": "B", "verdict": "bad"},
              "B": {"src_face": "B", "verdict": "bad"}
            }
          },
          "summary": {
            "good_count": 4,
            "total_cells": 12,
            "good_fraction": 0.333,
            "src_mismatch_count": 5  # slots where src_face != slot label
          }
        },
        ...
      ]
    }

The summary section enables quick triage and ranking by failure rate.
"""
from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl


XLSX_DEFAULT = Path(
    "/private/tmp/hybrid_overlays/overlay tool visual feedback  (1) copy.xlsx"
)
OUTPUT_DEFAULT = Path(__file__).resolve().parent.parent / (
    "tests/fixtures/overlay_visual_labels.json"
)

# Column conventions from the xlsx:
#   A quads:      cols 1-3 = red/green/blue outlines  → U/R/F quads
#   A rectified:  cols 4-6 = U/R/F face rectified
#   B quads:      cols 7-9 = purple/orange/yellow outlines → D/L/B quads
#   B rectified:  cols 10-12 = D/L/B face rectified
COLOR_TO_SLOT_A = {"red": "U", "green": "R", "blue": "F"}
COLOR_TO_SLOT_B = {"purple": "D", "orange": "L", "yellow": "B"}
SLOTS_A = ["U", "R", "F"]
SLOTS_B = ["D", "L", "B"]


def _normalize_verdict(raw: Optional[str]) -> Tuple[Optional[str], Optional[str]]:
    """Parse a cell value like 'src = U; good' or 'bad' or 'src = B; bad'.

    Returns (src_face, verdict) where:
      * src_face is "U"/"R"/"F"/"D"/"L"/"B" or None if not specified
      * verdict is "good" / "bad" / None
    """
    if raw is None:
        return None, None
    text = str(raw).strip().lower().replace("srd", "src")  # tolerate typo
    if not text:
        return None, None
    # Pattern: "src = X; good" or "src = X; bad"
    m = re.match(r"src\s*=\s*([a-z]+)\s*[;,]\s*(good|bad)", text)
    if m:
        return m.group(1).upper(), m.group(2)
    # Bare verdict: just "good" or "bad"
    if text in ("good", "bad"):
        return None, text
    # Unknown — preserve verbatim for debugging
    return None, text


def parse_xlsx(xlsx_path: Path) -> Dict[str, Any]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        raise ValueError(f"Expected at least 3 rows, got {len(rows)}")

    # rows[0] = section labels (unused after schema validation)
    # rows[1] = per-column header text
    # rows[2:] = data rows
    pairs: List[Dict[str, Any]] = []
    for raw_row in rows[2:]:
        if not raw_row or raw_row[0] is None:
            continue
        try:
            set_id_raw = raw_row[0]
            # May be "17.0" (Excel number) or "17"
            if isinstance(set_id_raw, (int, float)):
                set_id = str(int(set_id_raw))
            else:
                set_id = str(set_id_raw).strip().split(".")[0]
        except Exception:
            continue
        # Columns 1-3: A quads (red/green/blue → U/R/F)
        quad_A: Dict[str, str] = {}
        for i, slot in enumerate(SLOTS_A):
            _, verdict = _normalize_verdict(raw_row[1 + i])
            if verdict:
                quad_A[slot] = verdict
        # Columns 4-6: A rectified (U/R/F)
        rect_A: Dict[str, Dict[str, str]] = {}
        for i, slot in enumerate(SLOTS_A):
            src_face, verdict = _normalize_verdict(raw_row[4 + i])
            if verdict:
                rect_A[slot] = {"src_face": src_face, "verdict": verdict}
        # Columns 7-9: B quads (purple/orange/yellow → D/L/B)
        quad_B: Dict[str, str] = {}
        for i, slot in enumerate(SLOTS_B):
            _, verdict = _normalize_verdict(raw_row[7 + i])
            if verdict:
                quad_B[slot] = verdict
        # Columns 10-12: B rectified (D/L/B)
        rect_B: Dict[str, Dict[str, str]] = {}
        for i, slot in enumerate(SLOTS_B):
            src_face, verdict = _normalize_verdict(raw_row[10 + i])
            if verdict:
                rect_B[slot] = {"src_face": src_face, "verdict": verdict}

        # Summary metrics
        all_verdicts = (
            list(quad_A.values()) + list(quad_B.values())
            + [r["verdict"] for r in rect_A.values()]
            + [r["verdict"] for r in rect_B.values()]
        )
        good_count = sum(1 for v in all_verdicts if v == "good")
        total_cells = len(all_verdicts)
        src_mismatch_count = sum(
            1 for slot, rec in {**rect_A, **rect_B}.items()
            if rec.get("src_face") and rec["src_face"] != slot
        )

        pairs.append({
            "set_id": set_id,
            "quads": {"A": quad_A, "B": quad_B},
            "rectified": {"A": rect_A, "B": rect_B},
            "summary": {
                "good_count": good_count,
                "total_cells": total_cells,
                "good_fraction": (
                    round(good_count / total_cells, 4) if total_cells else 0.0
                ),
                "src_mismatch_count": src_mismatch_count,
            },
        })

    return {
        "source": str(xlsx_path),
        "captured_at": "2026-05-18",
        "pipeline_version_at_capture": "pre_PR_160",
        "notes": (
            "Per-slot human visual judgments from overlay-tool review. "
            "'good'/'bad' is the visual verdict; 'src_face' is the analyze_image "
            "center-face attribution for the rectified slot (when present in "
            "the cell text). 'src_face' != slot label indicates a slot/src "
            "mismatch — analyze_image classified a grid for face X but the "
            "proposer assigned it to slot Y. This was the dominant failure "
            "mode in the May 18 review."
        ),
        "color_to_slot": {"A": COLOR_TO_SLOT_A, "B": COLOR_TO_SLOT_B},
        "pairs": pairs,
    }


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--xlsx", default=str(XLSX_DEFAULT),
                    help=f"path to overlay xlsx (default: {XLSX_DEFAULT})")
    ap.add_argument("--out", default=str(OUTPUT_DEFAULT),
                    help=f"output json path (default: {OUTPUT_DEFAULT})")
    args = ap.parse_args()
    xlsx = Path(args.xlsx)
    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    data = parse_xlsx(xlsx)
    out.write_text(json.dumps(data, indent=2) + "\n")
    print(f"wrote {len(data['pairs'])} pairs to {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
